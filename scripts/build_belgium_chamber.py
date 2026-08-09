#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import time
from collections import Counter, defaultdict
from pathlib import Path

import requests
from openpyxl import load_workbook
from shapely.geometry import mapping, shape


SOURCES = {
    2024: {
        "results": (
            "https://elections.fgov.be/sites/2024.elections.fgov.be/files/inline-files/CK_CEListes_2024.xlsx",
            "45aef2aacca862f9338cc657efb5373c558c035b83cb79ebb632e73c8269e659",
        ),
        "turnout": (
            "https://elections.fgov.be/sites/2024.elections.fgov.be/files/inline-files/CK_TauxParticip_2024.xlsx",
            "748a06bab361ac493bb9d4a9e95b0e345e8523588bacef891b849658cb068bad",
        ),
    },
    2019: {
        "results": (
            "https://elections.fgov.be/sites/default/files/inline-files/CK_CEListes.xlsx",
            "64e5d8d0e4edc3072ace1a79cfa886157f755456f43da75bb37982a1908ac1da",
        ),
        "turnout": (
            "https://elections.fgov.be/sites/default/files/inline-files/CK_TauxParticip.xlsx",
            "e803fa445fc66401b0eaad52ae2a0efe16df7153b543d83769d8ff1eb11a5987",
        ),
    },
}
SOURCE_PAGE = {
    2024: "https://elections.fgov.be/elections-du-9-juin-2024-tableaux-des-resultats",
    2019: "https://elections.fgov.be/elections-du-26-mai-2019-tableaux-des-resultats",
}
BOUNDARY_URL = (
    "https://gisco-services.ec.europa.eu/distribution/v2/nuts/geojson/"
    "NUTS_RG_03M_2021_4326_LEVL_2.geojson"
)
BOUNDARY_SHA256 = "690c055cdaf583c5c2326dc8c36f2db9f60c42a256a498f8391d2d83fa89b5fb"

CONSTITUENCIES = {
    "C11002": ("Antwerp", "BE21"),
    "C21004": ("Brussels-Capital", "BE10"),
    "C24062": ("Flemish Brabant", "BE24"),
    "C25072": ("Walloon Brabant", "BE31"),
    "C31005": ("West Flanders", "BE25"),
    "C44021": ("East Flanders", "BE23"),
    "C53053": ("Hainaut", "BE32"),
    "C62063": ("Liège", "BE33"),
    "C71022": ("Limburg", "BE22"),
    "C81001": ("Luxembourg", "BE34"),
    "C92094": ("Namur", "BE35"),
}

FIELDS = (
    "district", "district_url", "distribution_url", "elected_member", "elected_party",
    "enrolment", "formal_votes", "informal_votes", "total_votes", "turnout_pct",
    "majority", "round_number", "row_type", "excluded_candidate", "excluded_party",
    "candidate", "candidate_party", "votes", "electorate_type", "constituency_code",
    "contest_status", "result_note", "district_seats",
)


def download(session: requests.Session, url: str, path: Path, refresh: bool) -> Path:
    if path.exists() and path.stat().st_size and not refresh:
        return path
    path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(5):
        try:
            response = session.get(url, timeout=600)
            response.raise_for_status()
            path.write_bytes(response.content)
            break
        except requests.RequestException:
            if attempt == 4:
                raise
            time.sleep(2 ** attempt)
    return path


def require_sha256(path: Path, expected: str) -> None:
    actual = hashlib.sha256(path.read_bytes()).hexdigest()
    if actual != expected:
        raise SystemExit(f"{path}: checksum changed; expected {expected}, found {actual}")


def party_name(value: object) -> str:
    text = str(value or "").strip()
    if "\n-\n" not in text:
        raise SystemExit(f"Belgian workbook has unexpected party header {text!r}")
    return text.split("\n-\n", 1)[1].strip()


def constituency_for_ins(value: object) -> str:
    code = str(int(value)).zfill(5)
    prefix = int(code[:2])
    if prefix == 21:
        return "C21004"
    if prefix in {23, 24}:
        return "C24062"
    if prefix == 25:
        return "C25072"
    return {
        "1": "C11002", "3": "C31005", "4": "C44021", "5": "C53053",
        "6": "C62063", "7": "C71022", "8": "C81001", "9": "C92094",
    }[code[0]]


def parse_turnout(path: Path) -> dict[str, dict[str, int]]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    totals = defaultdict(lambda: {"enrolment": 0, "total": 0, "informal": 0, "formal": 0})
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if values[0] is None:
            continue
        constituency = constituency_for_ins(values[0])
        enrolment = sum(int(value or 0) for value in values[3:7])
        total = sum(int(value or 0) for value in values[8:11])
        informal = int(values[14] or 0)
        formal = int(values[15] or 0)
        if total != formal + informal or total > enrolment:
            raise SystemExit(f"Belgium turnout row {values[0]} does not reconcile")
        totals[constituency]["enrolment"] += enrolment
        totals[constituency]["total"] += total
        totals[constituency]["informal"] += informal
        totals[constituency]["formal"] += formal
    if set(totals) != set(CONSTITUENCIES):
        raise SystemExit("Belgian turnout workbook does not cover all 11 constituencies")
    return dict(totals)


def dhondt(votes: dict[str, int], seat_count: int) -> dict[str, int]:
    eligible = {party: value for party, value in votes.items() if value >= sum(votes.values()) * 0.05}
    seats = Counter()
    for _ in range(seat_count):
        winner = max(eligible, key=lambda party: (eligible[party] / (seats[party] + 1), party))
        seats[winner] += 1
    return dict(seats)


def parse_results(path: Path, year: int) -> tuple[dict[str, dict[str, int]], Counter[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    results = {}
    parliament = Counter()
    for code in CONSTITUENCIES:
        sheet = workbook[code]
        parties = [party_name(sheet.cell(1, column).value) for column in range(2, sheet.max_column + 1)]
        votes = {party: int(sheet.cell(6, column).value or 0) for party, column in zip(parties, range(2, sheet.max_column + 1))}
        seat_sheet = workbook[f"Seat_{code}"]
        seat_parties = [party_name(seat_sheet.cell(1, column).value) for column in range(2, seat_sheet.max_column + 1)]
        seat_row = next(
            row for row in range(1, seat_sheet.max_row + 1)
            if str(seat_sheet.cell(row, 1).value or "").lower().startswith(("sièges", "sieges"))
        )
        seats = {
            party: int(seat_sheet.cell(seat_row, column).value or 0)
            for party, column in zip(seat_parties, range(2, seat_sheet.max_column + 1))
        }
        # The published 2024 East Flanders seat worksheet ends before the cd&v
        # column and therefore exposes only 18 of its statutory 20 seats. The
        # official list votes reproduce the missing two cd&v seats under D'Hondt.
        if year == 2024 and code == "C44021" and sum(seats.values()) == 18:
            seats = dhondt(votes, 20)
        parliament.update(seats)
        results[code] = {"votes": votes, "seats": seats}
    if sum(parliament.values()) != 150:
        raise SystemExit(f"Belgium: expected 150 Chamber seats, found {sum(parliament.values())}")
    return results, parliament


def build_boundaries(path: Path) -> dict[str, object]:
    source = json.loads(path.read_text(encoding="utf-8"))
    by_code = {feature["properties"]["NUTS_ID"]: feature for feature in source["features"]}
    features = []
    for _, (name, nuts_code) in CONSTITUENCIES.items():
        geometry = shape(by_code[nuts_code]["geometry"]).simplify(0.002, preserve_topology=True)
        features.append({
            "type": "Feature",
            "properties": {"district": name, "constituency_code": nuts_code, "electorate_type": "Belgium"},
            "geometry": mapping(geometry),
        })
    return {
        "type": "FeatureCollection", "name": "belgium_chamber_constituencies",
        "source": "Eurostat GISCO NUTS 2021; Belgium's Chamber constituencies correspond to provinces plus Brussels-Capital",
        "features": features,
    }


def build_year(year: int, result_path: Path, turnout_path: Path, out_dir: Path) -> None:
    results, parliament = parse_results(result_path, year)
    turnout = parse_turnout(turnout_path)
    # Six Flemish-Brabant facility municipalities can cast Chamber ballots for
    # Brussels lists. SPF's turnout workbook groups those ballots by counting
    # canton while its list-vote workbook assigns them to the legal constituency.
    # Preserve the exact joint totals and apportion only enrolment/invalid-ballot
    # metadata between the two displayed constituencies by their valid-vote share.
    brabant_codes = ("C21004", "C24062")
    joint = {
        key: sum(turnout[code][key] for code in brabant_codes)
        for key in ("enrolment", "total", "informal", "formal")
    }
    brabant_formal = {code: sum(results[code]["votes"].values()) for code in brabant_codes}
    if sum(brabant_formal.values()) != joint["formal"]:
        raise SystemExit(f"Belgium {year}: joint Brussels/Brabant valid votes do not reconcile")
    first = brabant_codes[0]
    allocated_informal = round(joint["informal"] * brabant_formal[first] / joint["formal"])
    informal_by_code = {first: allocated_informal, brabant_codes[1]: joint["informal"] - allocated_informal}
    total_by_code = {code: brabant_formal[code] + informal_by_code[code] for code in brabant_codes}
    allocated_enrolment = round(joint["enrolment"] * total_by_code[first] / joint["total"])
    enrolment_by_code = {first: allocated_enrolment, brabant_codes[1]: joint["enrolment"] - allocated_enrolment}
    for code in brabant_codes:
        turnout[code] = {
            "enrolment": enrolment_by_code[code], "formal": brabant_formal[code],
            "informal": informal_by_code[code], "total": total_by_code[code],
        }
    rows = []
    for code, (name, nuts_code) in CONSTITUENCIES.items():
        votes = results[code]["votes"]
        seats = results[code]["seats"]
        metrics = turnout[code]
        if sum(votes.values()) != metrics["formal"]:
            raise SystemExit(f"Belgium {year} {name}: party votes do not equal valid ballots")
        ranked = sorted(votes.items(), key=lambda item: (-item[1], item[0]))
        district_seats = sum(seats.values())
        base = {
            "district": name, "district_url": SOURCE_PAGE[year], "distribution_url": BOUNDARY_URL,
            "elected_member": ranked[0][0], "elected_party": ranked[0][0],
            "enrolment": metrics["enrolment"], "formal_votes": metrics["formal"],
            "informal_votes": metrics["informal"], "total_votes": metrics["total"],
            "turnout_pct": round(metrics["total"] / metrics["enrolment"] * 100, 2),
            "majority": ranked[0][1] - ranked[1][1], "round_number": 0,
            "row_type": "first", "excluded_candidate": "", "excluded_party": "",
            "electorate_type": "Belgium", "constituency_code": nuts_code,
            "contest_status": "official", "district_seats": district_seats,
            "result_note": (
                "Official SPF Interior Chamber returns; the map shows the locally leading list in each multi-member constituency."
                + (
                    " Brussels/Flemish-Brabant enrolment and invalid ballots are proportionally apportioned from their exact joint canton totals because facility-municipality ballots are assigned across the constituency boundary."
                    if code in brabant_codes else ""
                )
            ),
        }
        for party, party_votes in ranked:
            rows.append({**base, "candidate": party, "candidate_party": party, "votes": party_votes})
    output = out_dir / f"belgium_{year}_chamber_fpp.csv"
    with output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    print(f"Belgium {year}: wrote {len(rows)} rows, 11 constituencies and 150 seats")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Belgium 2024 and 2019 federal Chamber maps")
    parser.add_argument("--raw-dir", type=Path, default=Path("tmp/belgium_chamber"))
    parser.add_argument("--out-dir", type=Path, default=Path("data"))
    parser.add_argument("--refresh", action="store_true")
    args = parser.parse_args()
    session = requests.Session()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    boundary_path = download(session, BOUNDARY_URL, args.raw_dir / "nuts2_2021.geojson", args.refresh)
    require_sha256(boundary_path, BOUNDARY_SHA256)
    (args.out_dir / "belgium_chamber_boundaries.geojson").write_text(
        json.dumps(build_boundaries(boundary_path), ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    for year in (2024, 2019):
        paths = {}
        for key, (url, expected_hash) in SOURCES[year].items():
            path = download(session, url, args.raw_dir / f"belgium_{year}_{key}.xlsx", args.refresh)
            require_sha256(path, expected_hash)
            paths[key] = path
        build_year(year, paths["results"], paths["turnout"], args.out_dir)


if __name__ == "__main__":
    main()
