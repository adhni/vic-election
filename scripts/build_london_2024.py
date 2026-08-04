#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path

import openpyxl
import requests
from shapely.geometry import mapping, shape


RESULTS_URL = (
    "https://www.londonelects.org.uk/files/"
    "GLA%20Election%20May%202024%20Results%20Summary.xlsx"
)
RESULTS_SHA256 = "3984fde370e3f0e8c4960bb80e7a7adfffe951da770d19ddd0aff6ad8a1f0934"
BOUNDARIES_URL = (
    "https://gis.london.gov.uk/arcgis/rest/services/apps/webmap_context_layer/"
    "MapServer/7/query?where=1%3D1&outFields=lac_name&returnGeometry=true&outSR=4326&f=geojson"
)
RESULTS_PAGE = "https://www.londonelects.org.uk/im-voter/election-results/results-2024/"
UA = "Mozilla/5.0 (compatible; election-preference-explorer/0.1; +https://github.com/)"

FIELDS = [
    "district", "district_url", "distribution_url", "elected_member", "elected_party",
    "enrolment", "formal_votes", "informal_votes", "total_votes", "turnout_pct",
    "majority", "round_number", "row_type", "excluded_candidate", "excluded_party",
    "candidate", "candidate_party", "votes", "electorate_type", "constituency_code",
    "contest_status", "result_note",
]
MEMBER_FIELDS = ["seat_type", "seat_number", "constituency", "member", "party"]

AREA_NAMES = {
    "Barnet & Camden": "Barnet and Camden",
    "Brent & Harrow": "Brent and Harrow",
    "City & East": "City and East",
    "Croydon & Sutton": "Croydon and Sutton",
    "Havering & Redbridge": "Havering and Redbridge",
}

PARTIES = {
    "Animal Welfare Party - People, Animals, Environment": "Animal Welfare Party",
    "Conservative and Unionist Party": "Conservative",
    "Conservatives": "Conservative",
    "The Conservative Party Candidate": "Conservative",
    "Green Party": "Green Party",
    "The Green Party": "Green Party",
    "Labour and Co-operative Party": "Labour",
    "Labour Party": "Labour",
    "Liberal Democrat": "Liberal Democrats",
    "Liberal Democrats": "Liberal Democrats",
    "ReformUK": "Reform UK",
    "ReformUK – London Deserves Better": "Reform UK",
}


def download(session: requests.Session, url: str, path: Path, refresh: bool) -> Path:
    if path.exists() and not refresh:
        return path
    response = session.get(url, timeout=120)
    response.raise_for_status()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(response.content)
    return path


def require_sha256(path: Path, expected: str) -> None:
    actual = hashlib.sha256(path.read_bytes()).hexdigest()
    if actual != expected:
        raise SystemExit(f"{path}: source checksum changed ({actual})")


def area_name(value: object) -> str:
    text = str(value or "").strip()
    return AREA_NAMES.get(text, text)


def area_code(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return f"LON24-{slug}"


def party_name(value: object) -> str:
    text = str(value or "Independent").strip()
    if " - Independent" in text or text.endswith("Independent"):
        return "Independent"
    return PARTIES.get(text, text)


def result_row(
    district: str,
    candidates: list[tuple[str, str, int]],
    formal: int,
    informal: int,
    total: int,
    enrolment: int,
    note: str,
) -> list[dict[str, object]]:
    if sum(votes for _, _, votes in candidates) != formal:
        raise SystemExit(f"{district}: candidate/list votes do not equal good votes")
    if formal + informal != total:
        raise SystemExit(f"{district}: good plus rejected votes do not equal ballots counted")
    ranked = sorted(candidates, key=lambda item: (-item[2], item[0]))
    winner, winner_party, winner_votes = ranked[0]
    majority = winner_votes - ranked[1][2] if len(ranked) > 1 else winner_votes
    base = {
        "district": district,
        "district_url": RESULTS_PAGE,
        "distribution_url": RESULTS_PAGE,
        "elected_member": winner,
        "elected_party": winner_party,
        "enrolment": enrolment,
        "formal_votes": formal,
        "informal_votes": informal,
        "total_votes": total,
        "turnout_pct": round(total / enrolment * 100, 2) if enrolment else 0,
        "majority": majority,
        "round_number": 0,
        "row_type": "first",
        "excluded_candidate": "",
        "excluded_party": "",
        "electorate_type": "London",
        "constituency_code": area_code(district),
        "contest_status": "official",
        "result_note": note,
    }
    return [{**base, "candidate": candidate, "candidate_party": party, "votes": votes}
            for candidate, party, votes in candidates]


def split_mayor_header(value: object) -> tuple[str, str]:
    name, party = str(value).split("\n", 1)
    name = name.replace("Zoȇ", "Zoë")
    return name.strip(), party_name(party)


def build_mayor(workbook: openpyxl.Workbook) -> list[dict[str, object]]:
    sheet = workbook["1. Mayor of London"]
    headers = [sheet.cell(3, column).value for column in range(1, 24)]
    candidate_headers = [split_mayor_header(value) for value in headers[1:14]]
    output: list[dict[str, object]] = []
    for row in range(4, 18):
        district = area_name(sheet.cell(row, 1).value)
        candidates = [
            (name, party, int(sheet.cell(row, column).value or 0))
            for column, (name, party) in enumerate(candidate_headers, start=2)
        ]
        output.extend(result_row(
            district, candidates,
            int(sheet.cell(row, 15).value), int(sheet.cell(row, 20).value),
            int(sheet.cell(row, 21).value), int(sheet.cell(row, 22).value),
            "The Mayor is elected London-wide. The constituency map shows the locally leading mayoral candidate.",
        ))
    return output


def build_constituency_assembly(workbook: openpyxl.Workbook) -> list[dict[str, object]]:
    votes_sheet = workbook["2. Constituency Assembly Pt 1"]
    metadata_sheet = workbook["3. Constituency Assembly Pt 2"]
    grouped: dict[str, list[tuple[str, str, int]]] = {}
    for row in votes_sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        candidate, party, district, votes = row[:4]
        grouped.setdefault(area_name(district), []).append(
            (str(candidate).strip(), party_name(party), int(votes or 0))
        )
    metadata = {
        area_name(row[0]): row
        for row in metadata_sheet.iter_rows(min_row=4, max_row=17, values_only=True)
    }
    if len(grouped) != 14 or set(grouped) != set(metadata):
        raise SystemExit("Constituency Assembly vote and metadata areas do not match")
    output: list[dict[str, object]] = []
    for district in sorted(grouped):
        row = metadata[district]
        output.extend(result_row(
            district, grouped[district], int(row[1]), int(row[7]), int(row[8]), int(row[9]),
            "Each constituency elects one London Assembly member by first-past-the-post.",
        ))
    return output


def list_label(header: object) -> tuple[str, str]:
    text = str(header).replace("\n", " ").strip()
    if text == "FOX, Laurence":
        return "Laurence Fox", "Independent"
    independent = re.match(r"^(FOX|LONDON|ROMUALDO),\s+(.+?)\s+Independent$", text)
    if independent:
        return f"{independent.group(2).title()} {independent.group(1).title()}", "Independent"
    return text, party_name(text)


def build_london_wide(workbook: openpyxl.Workbook) -> list[dict[str, object]]:
    sheet = workbook["4. London-wide Assembly"]
    headers = [list_label(sheet.cell(3, column).value) for column in range(2, 17)]
    output: list[dict[str, object]] = []
    for row in range(4, 18):
        district = area_name(sheet.cell(row, 1).value)
        candidates = [
            (candidate, party, int(sheet.cell(row, column).value or 0))
            for column, (candidate, party) in enumerate(headers, start=2)
        ]
        output.extend(result_row(
            district, candidates,
            int(sheet.cell(row, 17).value), int(sheet.cell(row, 23).value),
            int(sheet.cell(row, 24).value), int(sheet.cell(row, 25).value),
            "London-wide list votes elect 11 Assembly members by modified D'Hondt. Constituencies show local list leaders only.",
        ))
    return output


def build_members(
    workbook: openpyxl.Workbook,
    constituency_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    members = []
    seen_districts = set()
    for row in constituency_rows:
        district = str(row["district"])
        if district in seen_districts:
            continue
        seen_districts.add(district)
        members.append({
            "seat_type": "constituency",
            "seat_number": "",
            "constituency": district,
            "member": row["elected_member"],
            "party": row["elected_party"],
        })

    sheet = workbook["7. London-wide Elected"]
    list_members = []
    for column in range(2, 17):
        raw_party = re.sub(r"\s+", " ", str(sheet.cell(1, column).value or "")).strip()
        party = party_name(raw_party)
        for row in range(2, 17):
            value = str(sheet.cell(row, column).value or "").strip()
            match = re.match(r"^(.*?) \(Elected In Round: (\d+)\)$", value)
            if not match:
                continue
            list_members.append({
                "seat_type": "london-wide",
                "seat_number": int(match.group(2)),
                "constituency": "",
                "member": match.group(1),
                "party": party,
            })
    if len(list_members) != 11:
        raise SystemExit(f"Expected 11 London-wide elected members, found {len(list_members)}")
    members.extend(sorted(list_members, key=lambda member: int(member["seat_number"])))
    return members


def build_boundaries(source: dict[str, object]) -> dict[str, object]:
    features = []
    for feature in source.get("features", []):
        district = area_name(feature.get("properties", {}).get("lac_name"))
        geometry = mapping(shape(feature["geometry"]).simplify(0.00035, preserve_topology=True))
        features.append({
            "type": "Feature",
            "properties": {
                "district": district,
                "constituency_code": area_code(district),
                "electorate_type": "London",
            },
            "geometry": geometry,
        })
    if len(features) != 14 or len({f["properties"]["district"] for f in features}) != 14:
        raise SystemExit("Expected 14 unique London Assembly constituency boundaries")
    features.sort(key=lambda feature: feature["properties"]["district"])
    return {"type": "FeatureCollection", "name": "london_2024_assembly_constituencies", "features": features}


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build official London Mayor and Assembly 2024 datasets")
    parser.add_argument("--raw-dir", type=Path, default=Path("tmp/london_2024"))
    parser.add_argument("--out-dir", type=Path, default=Path("data"))
    parser.add_argument("--refresh", action="store_true")
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update({"User-Agent": UA})
    workbook_path = download(session, RESULTS_URL, args.raw_dir / "results_summary.xlsx", args.refresh)
    boundary_path = download(session, BOUNDARIES_URL, args.raw_dir / "boundaries.geojson", args.refresh)
    require_sha256(workbook_path, RESULTS_SHA256)
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    outputs = {
        "london_2024_mayor_fpp.csv": build_mayor(workbook),
        "london_2024_assembly_constituency_fpp.csv": build_constituency_assembly(workbook),
        "london_2024_assembly_london_wide_fpp.csv": build_london_wide(workbook),
    }
    for filename, rows in outputs.items():
        write_csv(args.out_dir / filename, rows)
        print(f"Wrote {args.out_dir / filename} ({len(rows)} rows)")

    members = build_members(workbook, outputs["london_2024_assembly_constituency_fpp.csv"])
    member_path = args.out_dir / "london_2024_assembly_members.csv"
    with member_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=MEMBER_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(members)
    print(f"Wrote {member_path} ({len(members)} members)")

    boundaries = build_boundaries(json.loads(boundary_path.read_text(encoding="utf-8")))
    output_boundary = args.out_dir / "london_2024_assembly_constituencies.geojson"
    output_boundary.write_text(
        json.dumps(boundaries, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    print(f"Wrote {output_boundary} ({len(boundaries['features'])} features)")


if __name__ == "__main__":
    main()
